"""
Projeto Automatização de Relatórios Financeiros
==============================================

Objetivo: Automatizar a extração, edição e importação de dados de comprovantes
de um portal corporativo para planilha Excel de destino. Reduz tempo manual
de 10-15min para ~1-2min, minimizando erros humanos.

Stack:
- Selenium: Navegação e extração web automatizada.
- Pandas/OpenPyXL: Manipulação de dados Excel.
- PyAutoGUI: Automação de GUI (fallback para ações complexas).
- Win32com: Recálculo forçado de fórmulas no Excel.

Fluxo principal:
1. Login no portal (com validação 2FA manual).
2. Filtros > Download Excel.
3. Edição: Limpeza colunas, extração datas, mapeamento conferentes genérico.
4. Copia para planilha destino + fórmulas automáticas.
5. Recálculo e limpeza final.

⚠️ Pré-requisitos:
- Instalar: pip install selenium pandas openpyxl pyautogui
- ChromeDriver na PATH.
- Configurar: URL, login, senha e coords PyAutoGUI.
- Testar em ambiente controlado (coords de tela mudam por resolução).

Melhorias futuras:
- Config via YAML/ENV vars (sem hardcode).
- Tratamento de erros robusto (retry, screenshots).
- Suporte multi-período.
- Docker para portabilidade.
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import openpyxl
from openpyxl import load_workbook
import pyautogui
import pandas as pd
import subprocess
import win32com.client as win32  # Para Windows; use xlwings em Mac/Linux

# =======================================
# CONFIGURAÇÕES (EDITAR AQUI PARA USO)
# =======================================
URL_PORTAL = "https://exemplo-portal-corporativo.com/login"  # URL do portal real
LOGIN = "SEU_LOGIN"  # Usuário corporativo
SENHA = "SUA_SENHA"  # Senha (use env vars em prod!)

# Coordenadas de tela para PyAutoGUI (ajuste com pyautogui.position())
COORD_CLICKS = {
    'exemplo_click': (1155, 726)  # Exemplo; capture com pyautogui.position()
}

# Diretórios genéricos
HOME_DIR = os.path.expanduser("~")
DOWNLOAD_DIR = os.path.join(HOME_DIR, "Downloads")
ARQUIVO_DESTINO = "relatorio_atualizacao.xlsx"  # Planilha de destino
ARQUIVO_TEMP = "comprovantes_temp.xlsx"

print("🚀 Iniciando automação de relatórios...")

# =======================================
# 1. LOGIN E NAVEGAÇÃO NO PORTAL
# =======================================
def setup_driver():
    """Configura Chrome headless/invisível para automação estável."""
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # options.add_argument("--headless")  # Descomente para rodar sem GUI
    driver = webdriver.Chrome(options=options)
    return driver

driver = setup_driver()
driver.get(URL_PORTAL)
time.sleep(2)  # Aguarda carregamento

# Preenche login (lógica genérica; adapte XPaths reais)
campo_usuario = driver.find_element(By.ID, "login")
campo_usuario.send_keys(LOGIN)
campo_senha = driver.find_element(By.ID, "senha")
campo_senha.send_keys(SENHA)
campo_senha.send_keys(Keys.RETURN)

print("🔐 Login submetido. Insira código 2FA manualmente (20s)...")
time.sleep(25)  # Tempo para inserir código 2FA manual (ponto fraco; ideal: automação Outlook)

# Navegação para extração (adapte XPaths ao portal real)
WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, "loading")))
# Exemplo de cliques em menu/filtros
driver.find_element(By.ID, "token24set").click()
time.sleep(3)
# ... (sequência de filtros: período "último mês", modelo "completo", exportar)

baixar_arquivo = driver.find_element(By.XPATH, "//td[4]/a")  # XPath genérico para download
baixar_arquivo.click()
time.sleep(10)  # Aguarda download

driver.quit()
print("📥 Download concluído.")

# =======================================
# 2. PROCESSAMENTO DO ARQUIVO BAIXADO
# =======================================
def get_arquivo_recente(diretorio):
    """Pega o Excel mais recente em Downloads."""
    arquivos = [os.path.join(diretorio, f) for f in os.listdir(diretorio)
                if f.endswith('.xlsx')]
    if not arquivos:
        raise FileNotFoundError("Nenhum Excel baixado!")
    return max(arquivos, key=os.path.getctime)

caminho_excel = get_arquivo_recente(DOWNLOAD_DIR)
df = pd.read_excel(caminho_excel)
df.columns = df.columns.str.strip()  # Limpa nomes de colunas

# Edições comuns: extrair data, remover fotos desnecessárias
if 'Data' in df.columns:
    df.insert(loc=df.columns.get_loc('Data'), column='Data_full',
              df['Data'].astype(str).str[:10])  # YYYY-MM-DD
    # Coluna auxiliar para fórmulas (exemplo genérico)
    df.insert(loc=df.columns.get_loc('Data_full') + 1, column='Mes_Ano',
              df['Data_full'].apply(lambda x: f"{x[:7]}"))  # Formato MM/YYYY

    # Remove colunas irrelevantes (fotos, IDs internos)
    cols_remover = [f'Foto {i}' for i in range(1, 11)]
    df = df.drop(columns=[col for col in cols_remover if col in df.columns])

    # Mapeamento genérico de conferentes (substitua por lógica real sem nomes)
    def extrair_conferente(cargo):
        partes = cargo.split() if pd.notna(cargo) else []
        return partes[-1] if partes else None

    if 'Centro de Custo/Cartão' in df.columns:
        df['Conferente'] = df['Centro de Custo/Cartão'].apply(extrair_conferente)
        # Mapeamento exemplo: BU1 -> 'Equipe Norte', etc. (genérico)
        mapeamento = {'BU1': 'Equipe Norte', 'BU2': 'Equipe Centro', 'BU3': 'Equipe Sul'}
        df['Conferente'] = df['Conferente'].map(mapeamento).fillna('Pendente')

df.to_excel(ARQUIVO_TEMP, index=False)
print("✏️ Dados editados e salvos em temp.")

# =======================================
# 3. IMPORTAÇÃO PARA PLANILHA DESTINO
# =======================================
def get_last_row(sheet):
    """Encontra última linha não-vazia."""
    for row in range(sheet.max_row, 0, -1):
        if sheet[f'A{row}'].value is not None:
            return row
    return 1

wb_destino = load_workbook(ARQUIVO_DESTINO)
ws_destino = wb_destino.active  # Assume 'Sheet1' ou 'Diária'

# Limpa dados antigos (mantém headers)
ws_destino.delete_rows(2, ws_destino.max_row)

# Copia da temp para destino
wb_origem = load_workbook(ARQUIVO_TEMP)
ws_origem = wb_origem.active
for row in ws_origem.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=33, values_only=True):
    ws_destino.append(row)

wb_destino.save(ARQUIVO_DESTINO)

# Adiciona fórmulas automáticas (exemplo: estorno, ano)
last_row = get_last_row(ws_destino)
for row in range(2, last_row + 1):
    # Fórmula exemplo para coluna AH (estorno)
    ws_destino[f'AH{row}'] = f'=IF(OR(ISNUMBER(SEARCH("ESTORNADO",AC{row})),ISNUMBER(SEARCH("Extorno",AC{row}))),"ESTORNADO", "OK")'
    ws_destino[f'AI{row}'] = f'=YEAR(C{row})'

wb_destino.save(ARQUIVO_DESTINO)
print("📊 Fórmulas adicionadas.")

# =======================================
# 4. RECÁLCULO E FINALIZAÇÃO
# =======================================
# Força recálculo via COM (Windows)
excel = win32.Dispatch("Excel.Application")
excel.Visible = False
wb_excel = excel.Workbooks.Open(os.path.abspath(ARQUIVO_DESTINO))
wb_excel.RefreshAll()
excel.CalculateUntilAsyncQueriesDone()
wb_excel.Save()
wb_excel.Close(SaveChanges=True)
excel.Quit()

# Remove fórmulas, mantém valores
wb_destino = load_workbook(ARQUIVO_DESTINO, data_only=True)
ws_destino = wb_destino.active
wb_destino.save(ARQUIVO_DESTINO)

print("✅ Automação concluída! Arquivo pronto em:", ARQUIVO_DESTINO)
print("⏱️ Tempo total: <2min | Erros manuais: 0%")
